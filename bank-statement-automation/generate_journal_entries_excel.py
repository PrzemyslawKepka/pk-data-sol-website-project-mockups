"""
Generate a mock Excel file with journal entries in debit/credit format.
This creates accounting journal entries from bank transactions (two rows per transaction).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

# Output file
OUTPUT_FILE = "journal_entries_december_2024.xlsx"

# GL Account mappings (fictional)
GL_ACCOUNTS = {
    "bank": {"code": "10100", "name": "Bank Account - CHF"},
    "ar": {"code": "12000", "name": "Accounts Receivable"},
    "ap": {"code": "20000", "name": "Accounts Payable"},
    "interco_receivable": {"code": "13100", "name": "Intercompany Receivable"},
    "interco_payable": {"code": "21100", "name": "Intercompany Payable"},
    "bank_fees": {"code": "65100", "name": "Bank Charges & Fees"},
    "interest_income": {"code": "42000", "name": "Interest Income"},
    "fx_gain_loss": {"code": "43500", "name": "FX Gain/Loss"},
    "payroll": {"code": "62000", "name": "Payroll Expense"},
    "rent": {"code": "63000", "name": "Rent Expense"},
    "utilities": {"code": "63500", "name": "Utilities Expense"},
    "insurance": {"code": "64000", "name": "Insurance Expense"},
    "tax_payable": {"code": "22000", "name": "Tax Payable - VAT"},
    "professional_fees": {"code": "66000", "name": "Professional Services"},
}

def generate_transactions():
    """Generate mock transactions for journal entries."""

    transactions = []
    start_date = datetime(2024, 12, 1)

    # Transaction templates: (description, category, min_amount, max_amount, is_incoming)
    transaction_types = [
        ("Customer payment - INV-{}", "ar", 5000, 45000, True),
        ("Payment received - Order #{}", "ar", 1000, 12000, True),
        ("Intercompany transfer from TechFlow UK", "interco_receivable", 15000, 80000, True),
        ("Intercompany transfer from TechFlow DE", "interco_receivable", 20000, 60000, True),
        ("Interest income - December", "interest_income", 100, 800, True),
        ("FX Gain adjustment", "fx_gain_loss", 200, 1500, True),
        ("Supplier payment - Nordic Supply Ltd", "ap", 3000, 25000, False),
        ("Supplier payment - Alpine Systems AG", "ap", 2000, 18000, False),
        ("Supplier payment - Central Trading Co", "ap", 5000, 30000, False),
        ("Payroll transfer - December", "payroll", 80000, 120000, False),
        ("Bank charges - Monthly fee", "bank_fees", 35, 75, False),
        ("SWIFT transfer fee", "bank_fees", 20, 45, False),
        ("Intercompany payment to TechFlow US", "interco_payable", 25000, 70000, False),
        ("Rent payment - Office Q4", "rent", 8000, 12000, False),
        ("Insurance premium - Annual", "insurance", 2500, 5000, False),
        ("Utility payment - December", "utilities", 800, 2000, False),
        ("Tax payment - VAT Q4", "tax_payable", 15000, 45000, False),
        ("Audit fee - Year end", "professional_fees", 5000, 12000, False),
    ]

    journal_number = 1001

    for day in range(1, 32):
        current_date = start_date + timedelta(days=day-1)

        # Skip weekends occasionally
        if current_date.weekday() >= 5 and random.random() > 0.3:
            continue

        # 1-4 transactions per day
        num_transactions = random.randint(1, 4)

        for _ in range(num_transactions):
            template = random.choice(transaction_types)
            description, category, min_amt, max_amt, is_incoming = template

            amount = round(random.uniform(min_amt, max_amt), 2)

            # Format description with random numbers if needed
            if "{}" in description:
                desc = description.format(random.randint(1000, 9999))
            else:
                desc = description

            transactions.append({
                "journal_no": f"JE-{journal_number}",
                "date": current_date.strftime("%d.%m.%Y"),
                "description": desc,
                "category": category,
                "amount": amount,
                "is_incoming": is_incoming
            })

            journal_number += 1

    return transactions

def create_excel():
    """Create the Excel file with journal entries in debit/credit format."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal Entries"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    currency_format = '#,##0.00'

    # Headers
    headers = [
        "Journal No",
        "Date",
        "GL Account",
        "Account Name",
        "Description",
        "Debit",
        "Credit",
        "Memo"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Generate transactions
    transactions = generate_transactions()

    row = 2
    total_debits = 0
    total_credits = 0

    for txn in transactions:
        # For each transaction, we create TWO rows (double-entry bookkeeping)
        # Row 1: Bank account entry
        # Row 2: Offset account entry

        bank_account = GL_ACCOUNTS["bank"]
        offset_account = GL_ACCOUNTS[txn["category"]]

        if txn["is_incoming"]:
            # Money coming in: Debit Bank, Credit offset
            # Row 1: Debit Bank
            ws.cell(row=row, column=1, value=txn["journal_no"]).alignment = center_align
            ws.cell(row=row, column=2, value=txn["date"]).alignment = center_align
            ws.cell(row=row, column=3, value=bank_account["code"]).alignment = center_align
            ws.cell(row=row, column=4, value=bank_account["name"])
            ws.cell(row=row, column=5, value=txn["description"])
            debit_cell = ws.cell(row=row, column=6, value=txn["amount"])
            debit_cell.number_format = currency_format
            debit_cell.alignment = right_align
            ws.cell(row=row, column=7, value="")  # Credit is empty
            ws.cell(row=row, column=8, value=f"Bank receipt - {txn['date']}")

            total_debits += txn["amount"]

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

            row += 1

            # Row 2: Credit offset account
            ws.cell(row=row, column=1, value=txn["journal_no"]).alignment = center_align
            ws.cell(row=row, column=2, value=txn["date"]).alignment = center_align
            ws.cell(row=row, column=3, value=offset_account["code"]).alignment = center_align
            ws.cell(row=row, column=4, value=offset_account["name"])
            ws.cell(row=row, column=5, value=txn["description"])
            ws.cell(row=row, column=6, value="")  # Debit is empty
            credit_cell = ws.cell(row=row, column=7, value=txn["amount"])
            credit_cell.number_format = currency_format
            credit_cell.alignment = right_align
            ws.cell(row=row, column=8, value="Offset entry")

            total_credits += txn["amount"]

        else:
            # Money going out: Credit Bank, Debit offset
            # Row 1: Credit Bank
            ws.cell(row=row, column=1, value=txn["journal_no"]).alignment = center_align
            ws.cell(row=row, column=2, value=txn["date"]).alignment = center_align
            ws.cell(row=row, column=3, value=bank_account["code"]).alignment = center_align
            ws.cell(row=row, column=4, value=bank_account["name"])
            ws.cell(row=row, column=5, value=txn["description"])
            ws.cell(row=row, column=6, value="")  # Debit is empty
            credit_cell = ws.cell(row=row, column=7, value=txn["amount"])
            credit_cell.number_format = currency_format
            credit_cell.alignment = right_align
            ws.cell(row=row, column=8, value=f"Bank payment - {txn['date']}")

            total_credits += txn["amount"]

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

            row += 1

            # Row 2: Debit offset account
            ws.cell(row=row, column=1, value=txn["journal_no"]).alignment = center_align
            ws.cell(row=row, column=2, value=txn["date"]).alignment = center_align
            ws.cell(row=row, column=3, value=offset_account["code"]).alignment = center_align
            ws.cell(row=row, column=4, value=offset_account["name"])
            ws.cell(row=row, column=5, value=txn["description"])
            debit_cell = ws.cell(row=row, column=6, value=txn["amount"])
            debit_cell.number_format = currency_format
            debit_cell.alignment = right_align
            ws.cell(row=row, column=7, value="")  # Credit is empty
            ws.cell(row=row, column=8, value="Offset entry")

            total_debits += txn["amount"]

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

        row += 1

    # Add totals row
    row += 1
    ws.cell(row=row, column=5, value="TOTALS").font = Font(bold=True)
    total_debit_cell = ws.cell(row=row, column=6, value=total_debits)
    total_debit_cell.number_format = currency_format
    total_debit_cell.font = Font(bold=True)
    total_debit_cell.alignment = right_align

    total_credit_cell = ws.cell(row=row, column=7, value=total_credits)
    total_credit_cell.number_format = currency_format
    total_credit_cell.font = Font(bold=True)
    total_credit_cell.alignment = right_align

    # Adjust column widths
    column_widths = [12, 12, 12, 25, 40, 15, 15, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    wb.save(OUTPUT_FILE)
    print(f"Excel file generated: {OUTPUT_FILE}")
    print(f"Total journal entries: {len(transactions)}")
    print(f"Total rows (debit + credit lines): {(row - 3)}")
    print(f"Total Debits: CHF {total_debits:,.2f}")
    print(f"Total Credits: CHF {total_credits:,.2f}")
    print(f"Balance check: {'BALANCED' if abs(total_debits - total_credits) < 0.01 else 'UNBALANCED'}")

if __name__ == "__main__":
    create_excel()
